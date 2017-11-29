import openpyxl

wb = openpyxl.load_workbook('ABBREV.xlsx')
old_sheet = wb.active
new_sheet = wb.create_sheet(index=0, title="Clean")

foods = set()

nrows = old_sheet.max_row
ncols = old_sheet.max_column

ridx = 1;


for i in range(nrows):
    desc = old_sheet.cell(row = i+1, column = 2).value.lower()
    parts = desc.split(',')
    food = parts[0].strip()

    if(food not in foods):
        foods.add(food)
        for j in range(ncols):
            new_sheet.cell(row = ridx, column = j+1).value = old_sheet.cell(row = i+1, column = j+1).value
        new_sheet.cell(row = ridx, column = 2).value = food
        ridx += 1

wb.save('ABBREV_clean.xlsx')
